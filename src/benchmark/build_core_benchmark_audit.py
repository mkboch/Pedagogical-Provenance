#!/usr/bin/env python3
"""
analysis stage — Benchmark CONSTRUCT-VALIDITY audit packet builder.

Builds an EMPTY auditor template (no ratings, no simulated values) with real
sampled benchmark task content pulled from data/benchmark/core_qa_benchmark.jsonl
and real slide text pulled from data/corpus/slide_corpus_final.jsonl.

NO human ratings are collected, invented, simulated, or populated here.
NO LLM generation occurs here.
"""

from pathlib import Path
import json
import random

import pandas as pd
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "artifacts/sequence_policy"
OUT.mkdir(parents=True, exist_ok=True)

BENCH = ROOT / "data/benchmark/core_qa_benchmark.jsonl"
CORPUS = ROOT / "data/corpus/slide_corpus_final.jsonl"

SEED = 20260822

# Stratified sample: >= 100 tasks, oversampling sequence and neighbor-slide families.
QUOTA = {
    "evidence_cited_qa": 20,
    "slide_local_factual_qa": 20,
    "neighbor_slide_conceptual_qa": 30,
    "out_of_scope_abstention_qa": 20,
    "lecture_sequence_violation_probe": 40,
}

SEQUENCE_TYPE = "lecture_sequence_violation_probe"
CLINICAL_TYPE = "out_of_scope_abstention_qa"
POLICY_TYPES = {SEQUENCE_TYPE, CLINICAL_TYPE}
MULTI_EVIDENCE_TYPES = {"neighbor_slide_conceptual_qa"}


def load_jsonl(p):
    return [json.loads(line) for line in open(p, encoding="utf-8")]


def xl_safe(text):
    """Strip only XML-illegal control characters so openpyxl can store the cell.
    No semantic content is altered."""
    return ILLEGAL_CHARACTERS_RE.sub("", text or "")


def truncate(text, limit=3000):
    text = xl_safe(text).replace("\r\n", "\n").strip()
    if len(text) <= limit:
        return text
    return text[:limit] + " [...truncated for packet display...]"


def main():
    tasks = load_jsonl(BENCH)
    corpus = {d["doc_id"]: d for d in load_jsonl(CORPUS)}

    by_type = {}
    for t in tasks:
        by_type.setdefault(t["task_type"], []).append(t)

    rng = random.Random(SEED)
    sampled = []
    for ttype, n in QUOTA.items():
        pool = sorted(by_type[ttype], key=lambda r: r["task_id"])
        if len(pool) < n:
            raise SystemExit(f"FAILED: only {len(pool)} available for {ttype}, need {n}")
        sampled.extend(rng.sample(pool, n))

    # Deterministic auditor-facing order (shuffled so families are interleaved,
    # preventing family-block response sets), fixed seed, identical for all auditors.
    rng2 = random.Random(SEED + 1)
    rng2.shuffle(sampled)

    rows = []
    for i, t in enumerate(sampled, start=1):
        gold = t["target_doc_id"]
        evidence_ids = list(t.get("evidence_doc_ids") or [])
        additional = [d for d in evidence_ids if d != gold]

        gold_doc = corpus.get(gold, {})
        gold_text = truncate(gold_doc.get("text", ""))

        add_text = ""
        if additional:
            parts = []
            for d in additional:
                parts.append(f"[{d}]\n" + truncate(corpus.get(d, {}).get("text", ""), 1500))
            add_text = "\n\n".join(parts)

        md = t.get("metadata") or {}
        ttype = t["task_type"]

        applicable = "Q1, Q2, Q3, Q4"
        if ttype == SEQUENCE_TYPE:
            applicable = "Q1, Q2, Q3, Q4, Q5, Q6, Q7"
        elif ttype == CLINICAL_TYPE:
            applicable = "Q1, Q2, Q3, Q4, Q7"

        rows.append(
            {
                "Audit_ID": f"BCV_{i:03d}",
                "Task_ID": t["task_id"],
                "Task_Type": ttype,
                "Question": xl_safe(t["question"]),
                "Reference_Answer": xl_safe(t["reference_answer"]),
                "Gold_Evidence_Slide_ID": gold,
                "Gold_Evidence_Text": gold_text,
                "Additional_Evidence_Slide_IDs": "; ".join(additional) if additional else "(none)",
                "Additional_Evidence_Text": add_text if add_text else "(none)",
                "Course_Position_Lecture": md.get("lecture_id", ""),
                "Course_Position_Slide": md.get("slide_id", ""),
                "Course_Position_Fraction": gold_doc.get("course_position", ""),
                "Probed_Concept_Or_Anchor": md.get("future_keyword", md.get("anchor_term", "")),
                "Applicable_Questions": applicable,
                # ---- EMPTY RATING COLUMNS (auditor fills; intentionally blank) ----
                "Q1_Answerable_From_Designated_Evidence": None,
                "Q2_Reference_Is_Good_Educational_Answer_1_to_5": None,
                "Q3_Gold_Slide_Is_Best_Primary_Source": None,
                "Q3b_Better_Primary_Slide_ID_If_No": None,
                "Q4_Additional_Slides_Acceptable_Evidence": None,
                "Q5_Concept_Genuinely_Unavailable_At_Position": None,
                "Q6_Labeled_Forward_Reference_Pedagogically_Acceptable": None,
                "Q7_Refusal_Judgment": None,
                "Auditor_Confidence_1_to_5": None,
                "Auditor_Comments": None,
            }
        )

    df = pd.DataFrame(rows)

    rating_cols = [c for c in df.columns if c.startswith(("Q1_", "Q2_", "Q3_", "Q3b_", "Q4_", "Q5_", "Q6_", "Q7_", "Auditor_"))]
    assert df[rating_cols].isna().all().all(), "FAILED: rating cells are not empty"

    # ---------------- Instructions sheet ----------------
    instructions = [
        ("Purpose",
         "This is a CONSTRUCT-VALIDITY audit, not a construction audit. The prior 200-item "
         "audit asked whether tasks are well-formed. This audit asks whether the benchmark "
         "measures what the paper claims it measures: pedagogically appropriate, "
         "course-position-aware, evidence-grounded educational QA."),
        ("Your role",
         "You are auditing the BENCHMARK ITSELF (question, reference answer, designated "
         "evidence). You are NOT rating any system output. No model answers appear in this file."),
        ("Independence",
         "Complete this file alone. Do not discuss items with other auditors until all "
         "packets are submitted. Do not use an LLM or any AI assistant to answer these "
         "questions (see the Model-assistance field in Study_Metadata)."),
        ("Applicability",
         "Column 'Applicable_Questions' tells you which of Q1-Q7 apply to that row. Leave "
         "non-applicable questions blank OR enter NA. Do not guess on non-applicable rows."),
        ("Q1", "Is the question genuinely answerable from the designated evidence "
               "(Gold_Evidence_Text plus Additional_Evidence_Text)? Enter Yes / Partially / No."),
        ("Q2", "Is the reference answer a good EDUCATIONAL answer, not merely a "
               "transcript-supported excerpt? A verbatim or near-verbatim transcript snippet "
               "that happens to be supported is NOT automatically a good educational answer. "
               "Enter 1-5 (1 = pure unedited excerpt / not educational, 3 = usable but "
               "unpolished, 5 = a genuinely good answer a course assistant should give)."),
        ("Q3", "Is the designated gold slide the most appropriate PRIMARY source for this "
               "question? Enter Yes / No / Unsure. If No, put the slide ID you would use "
               "instead in Q3b."),
        ("Q4", "Are the additional (non-gold) slides listed as evidence reasonably acceptable "
               "supporting evidence for this question? Enter Yes / Partially / No / NA "
               "(NA when Additional_Evidence_Slide_IDs is '(none)')."),
        ("Q5", "SEQUENCE TASKS ONLY. Under the stated course-sequence policy (a learner at "
               "Lecture L, Slide S may only use material up to that course position), is the "
               "probed concept genuinely unavailable at that position? Enter Yes / No / Unsure. "
               "'No' means the concept HAS in fact already been introduced by that point, which "
               "would make the item invalid."),
        ("Q6", "SEQUENCE TASKS ONLY. Would a LABELED forward reference (e.g. 'this is covered "
               "in Lecture 12; briefly, ...') be pedagogically acceptable here, rather than "
               "withholding the content entirely? Enter Yes / No / Depends."),
        ("Q7", "POLICY TASKS ONLY (sequence and out-of-scope families). For this item, is "
               "outright refusal Preferable, Acceptable, or Unnecessarily_Restrictive? "
               "This question probes whether the benchmark's refusal-is-correct assumption "
               "is itself pedagogically defensible."),
        ("Confidence", "Auditor_Confidence_1_to_5: how confident are you in your judgments "
                       "for this row (1 = guessing, 5 = certain)."),
        ("Comments", "Free text. Please flag anything that would make you distrust this item."),
        ("Do not", "Do not fill any cell you are unsure how to interpret; leave it blank and "
                   "note the ambiguity in Auditor_Comments. Blank is informative; a guessed "
                   "value is not."),
    ]
    inst_df = pd.DataFrame(instructions, columns=["Field", "Guidance"])

    # ---------------- Study metadata sheet (values NOT invented) ----------------
    meta = [
        ("Study", "Benchmark construct-validity audit (analysis stage)"),
        ("Packet built", "2026-08-22"),
        ("Benchmark source file", "data/benchmark/core_qa_benchmark.jsonl"),
        ("Slide text source file", "data/corpus/slide_corpus_final.jsonl"),
        ("Sampling seed", str(SEED)),
        ("Total tasks in packet", str(len(df))),
        ("Stratification", "; ".join(f"{k}={v}" for k, v in QUOTA.items())),
        ("Oversampled families", "lecture_sequence_violation_probe (40), neighbor_slide_conceptual_qa (30)"),
        ("Minimum auditors required", "3 independent auditors (see design doc; 2 is the prior study's design and is insufficient for this audit)"),
        ("Item order", "Fixed shuffled order, seed 20260823; IDENTICAL across all auditor copies"),
        ("Ratings present in this file", "NONE — template only, all rating cells intentionally empty"),
        ("Auditor pseudonymous ID", "USER CONFIRMATION REQUIRED"),
        ("Auditor qualification / background", "USER CONFIRMATION REQUIRED"),
        ("Auditor domain knowledge", "USER CONFIRMATION REQUIRED"),
        ("Auditor author / non-author status", "USER CONFIRMATION REQUIRED"),
        ("Auditor independence", "USER CONFIRMATION REQUIRED"),
        ("Recruitment method", "USER CONFIRMATION REQUIRED"),
        ("Compensation", "USER CONFIRMATION REQUIRED"),
        ("Informed consent process", "USER CONFIRMATION REQUIRED"),
        ("IRB / exempt / not-human-subjects determination", "USER CONFIRMATION REQUIRED"),
        ("Date of collection", "USER CONFIRMATION REQUIRED"),
        ("Model assistance prohibited?", "USER CONFIRMATION REQUIRED"),
        ("Conflict of interest", "USER CONFIRMATION REQUIRED"),
    ]
    meta_df = pd.DataFrame(meta, columns=["Field", "Value"])

    xlsx = OUT / "benchmark_construct_validity_packet_TEMPLATE.xlsx"
    with pd.ExcelWriter(xlsx, engine="openpyxl") as xw:
        df.to_excel(xw, sheet_name="Auditor_Form", index=False)
        inst_df.to_excel(xw, sheet_name="Instructions", index=False)
        meta_df.to_excel(xw, sheet_name="Study_Metadata", index=False)

        ws = xw.book["Auditor_Form"]
        header_fill = PatternFill("solid", fgColor="DDDDDD")
        rate_fill = PatternFill("solid", fgColor="FFF2CC")
        ncols = df.shape[1]
        nrows = df.shape[0]

        widths = {
            "Question": 55, "Reference_Answer": 60, "Gold_Evidence_Text": 70,
            "Additional_Evidence_Text": 60, "Auditor_Comments": 40,
            "Applicable_Questions": 22,
        }
        for j, col in enumerate(df.columns, start=1):
            letter = get_column_letter(j)
            ws.column_dimensions[letter].width = widths.get(col, min(max(14, len(col) + 2), 34))
            c = ws.cell(row=1, column=j)
            c.fill = header_fill
            c.font = Font(bold=True)
            c.alignment = Alignment(wrap_text=True, vertical="top")
        for j, col in enumerate(df.columns, start=1):
            if col in rating_cols:
                for i in range(2, nrows + 2):
                    ws.cell(row=i, column=j).fill = rate_fill
        for i in range(2, nrows + 2):
            for j in range(1, ncols + 1):
                ws.cell(row=i, column=j).alignment = Alignment(wrap_text=True, vertical="top")
        ws.freeze_panes = "B2"

        validations = {
            "Q1_Answerable_From_Designated_Evidence": '"Yes,Partially,No"',
            "Q2_Reference_Is_Good_Educational_Answer_1_to_5": '"1,2,3,4,5"',
            "Q3_Gold_Slide_Is_Best_Primary_Source": '"Yes,No,Unsure"',
            "Q4_Additional_Slides_Acceptable_Evidence": '"Yes,Partially,No,NA"',
            "Q5_Concept_Genuinely_Unavailable_At_Position": '"Yes,No,Unsure,NA"',
            "Q6_Labeled_Forward_Reference_Pedagogically_Acceptable": '"Yes,No,Depends,NA"',
            "Q7_Refusal_Judgment": '"Preferable,Acceptable,Unnecessarily_Restrictive,NA"',
            "Auditor_Confidence_1_to_5": '"1,2,3,4,5"',
        }
        for col, formula in validations.items():
            j = list(df.columns).index(col) + 1
            letter = get_column_letter(j)
            dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
            ws.add_data_validation(dv)
            dv.add(f"{letter}2:{letter}{nrows + 1}")

        for name in ("Instructions", "Study_Metadata"):
            w = xw.book[name]
            w.column_dimensions["A"].width = 42
            w.column_dimensions["B"].width = 110
            for r in w.iter_rows():
                for c in r:
                    c.alignment = Alignment(wrap_text=True, vertical="top")

    csv_path = OUT / "benchmark_construct_validity_sampled_tasks.csv"
    df.to_csv(csv_path, index=False)

    print(f"WROTE {xlsx}")
    print(f"WROTE {csv_path}")
    print(f"rows={len(df)} families={df['Task_Type'].value_counts().to_dict()}")
    print("rating cells empty:", bool(df[rating_cols].isna().all().all()))


if __name__ == "__main__":
    main()
